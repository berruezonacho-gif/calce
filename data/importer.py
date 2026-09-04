"""
data/importer.py — Importar movimientos de flujo de caja desde Excel o CSV.

El tesorero suele tener sus cobranzas y pagos en una planilla. Este módulo lee
un archivo (xlsx o csv) y lo convierte en movimientos para la proyección.

Formato flexible: detecta columnas por nombre (fecha, concepto/detalle, monto o
entrada/salida). Acepta variantes comunes en español. Los montos pueden venir:
  - en una columna con signo (positivo=entra, negativo=sale), o
  - en dos columnas separadas (entradas / salidas / débito / crédito).

No requiere pandas: usa openpyxl para xlsx y el csv de la stdlib, para mantener
liviana la dependencia.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime


# Nombres de columna reconocidos (en minúscula, sin acentos)
_DATE_COLS = {"fecha", "date", "dia", "vencimiento", "fecha de pago", "f. pago"}
_LABEL_COLS = {"concepto", "detalle", "descripcion", "referencia", "nombre", "item", "glosa"}
_AMOUNT_COLS = {"monto", "importe", "amount", "valor", "total", "neto"}
_IN_COLS = {"entrada", "entradas", "ingreso", "ingresos", "credito", "credito", "haber", "cobros", "cobranza", "cobranzas"}
_OUT_COLS = {"salida", "salidas", "egreso", "egresos", "debito", "debito", "debe", "pagos", "pago"}
_REC_COLS = {"recurrencia", "recurrence", "repeticion", "frecuencia", "periodicidad"}
_ACCOUNT_COLS = {"cuenta", "account", "banco", "caja", "cuenta bancaria"}
_MEDIO_COLS = {"medio", "medio de pago", "medio de cobro", "forma de pago", "instrumento", "metodo"}


def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    for a, b in [("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u")]:
        s = s.replace(a, b)
    return s


def _parse_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):  # date
        try:
            return value.isoformat()[:10]
        except Exception:
            pass
    s = str(value).strip()
    # dd/mm/aaaa o dd-mm-aaaa
    m = re.match(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})", s)
    if m:
        d, mo, y = m.groups()
        y = int(y)
        if y < 100:
            y += 2000
        try:
            return datetime(y, int(mo), int(d)).date().isoformat()
        except ValueError:
            return None
    # aaaa-mm-dd
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return s[:10]
    return None


def _parse_amount(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    # Sacar símbolo de moneda y espacios
    s = s.replace("$", "").replace(" ", "").replace("ARS", "")
    neg = s.startswith("(") and s.endswith(")")  # contabilidad: (123) = negativo
    s = s.strip("()")
    # Formato argentino: 1.234.567,89 → 1234567.89
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        val = float(s)
        return -val if neg else val
    except ValueError:
        return None


def _rows_from_xlsx(content: bytes) -> list[list]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl no está instalado. Ejecutá: pip install openpyxl")
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _rows_from_csv(content: bytes) -> list[list]:
    text = content.decode("utf-8-sig", errors="replace")
    # Detectar separador (coma o punto y coma)
    sample = text[:2000]
    sep = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=sep)
    return [row for row in reader]


def parse_movements(content: bytes, filename: str = "") -> dict:
    """Convierte el contenido de un archivo en movimientos de cash flow.

    Devuelve {"ok", "movements": [...], "skipped": n, "columns": {...}}.
    """
    name = (filename or "").lower()
    try:
        if name.endswith(".csv") or (b"," in content[:200] and not name.endswith((".xlsx", ".xls"))):
            rows = _rows_from_csv(content)
        else:
            rows = _rows_from_xlsx(content)
    except Exception as e:
        return {"ok": False, "error": f"No se pudo leer el archivo: {e}", "movements": []}

    if not rows or len(rows) < 2:
        return {"ok": False, "error": "El archivo está vacío o no tiene datos.", "movements": []}

    # Detectar la fila de encabezados (la primera con al menos 2 celdas de texto)
    header_idx = 0
    for i, row in enumerate(rows[:5]):
        texts = [c for c in row if isinstance(c, str) and c.strip()]
        if len(texts) >= 2:
            header_idx = i
            break
    headers = [_norm(str(c)) for c in rows[header_idx]]

    # Mapear columnas
    col_date = col_label = col_amount = col_in = col_out = col_rec = None
    col_account = col_medio = None
    for i, h in enumerate(headers):
        if col_date is None and h in _DATE_COLS: col_date = i
        if col_label is None and h in _LABEL_COLS: col_label = i
        if col_amount is None and h in _AMOUNT_COLS: col_amount = i
        if col_in is None and h in _IN_COLS: col_in = i
        if col_out is None and h in _OUT_COLS: col_out = i
        if col_rec is None and h in _REC_COLS: col_rec = i
        if col_account is None and h in _ACCOUNT_COLS: col_account = i
        if col_medio is None and h in _MEDIO_COLS: col_medio = i

    if col_date is None:
        return {"ok": False, "error": "No se encontró una columna de fecha. Asegurate de que tenga un encabezado como 'Fecha'.", "movements": []}
    if col_amount is None and col_in is None and col_out is None:
        return {"ok": False, "error": "No se encontró columna de monto. Usá 'Monto' (con signo) o 'Entradas'/'Salidas'.", "movements": []}

    movements = []
    skipped = 0
    for row in rows[header_idx + 1:]:
        if not row or all(c is None or c == "" for c in row):
            continue
        d = _parse_date(row[col_date]) if col_date < len(row) else None
        if not d:
            skipped += 1
            continue
        amount = None
        if col_amount is not None and col_amount < len(row):
            amount = _parse_amount(row[col_amount])
        if amount is None and (col_in is not None or col_out is not None):
            inflow = _parse_amount(row[col_in]) if (col_in is not None and col_in < len(row)) else 0
            outflow = _parse_amount(row[col_out]) if (col_out is not None and col_out < len(row)) else 0
            amount = (inflow or 0) - (outflow or 0)
        if amount is None or amount == 0:
            skipped += 1
            continue
        label = ""
        if col_label is not None and col_label < len(row) and row[col_label]:
            label = str(row[col_label]).strip()
        rec = "none"
        if col_rec is not None and col_rec < len(row) and row[col_rec]:
            rv = _norm(str(row[col_rec]))
            if rv in ("weekly", "semanal", "semana"): rec = "weekly"
            elif rv in ("quincenal", "quincena", "biweekly"): rec = "quincenal"
            elif rv in ("monthly", "mensual", "mes"): rec = "monthly"
            elif rv in ("quarterly", "trimestral", "trimestre"): rec = "quarterly"
            elif rv in ("daily", "diario", "dia"): rec = "daily"
        # Texto de cuenta y medio (el frontend los resuelve contra sus cuentas)
        account_text = ""
        if col_account is not None and col_account < len(row) and row[col_account]:
            account_text = str(row[col_account]).strip()
        medio = ""
        if col_medio is not None and col_medio < len(row) and row[col_medio]:
            mv = _norm(str(row[col_medio]))
            if "transfer" in mv: medio = "transferencia"
            elif "efectivo" in mv or "cash" in mv: medio = "efectivo"
            elif "cheque" in mv or "echeq" in mv: medio = "cheque"
            elif "tarjeta" in mv or "card" in mv: medio = "tarjeta"
        movements.append({
            "date": d,
            "amount": amount,
            "label": label,
            "recurrence": rec,
            "account_text": account_text,
            "medio": medio,
        })

    return {
        "ok": True,
        "movements": movements,
        "count": len(movements),
        "skipped": skipped,
        "columns": {
            "date": headers[col_date] if col_date is not None else None,
            "amount": headers[col_amount] if col_amount is not None else None,
            "in": headers[col_in] if col_in is not None else None,
            "out": headers[col_out] if col_out is not None else None,
        },
    }


# ══════════════════════════════════════════════════════════════════
# Import del Libro IVA Compras de AFIP/ARCA ("Mis Comprobantes")
# ══════════════════════════════════════════════════════════════════
# El export de AFIP trae, por cada comprobante de compra: fecha, tipo,
# punto de venta, número, CUIT y denominación del proveedor, y los
# importes discriminados (Neto Gravado, No Gravado, Exento, IVA, Total).
# De un solo archivo obtenemos: (a) el directorio de proveedores (CUIT
# únicos) y (b) las facturas a pagar con IVA discriminado — sin estimar.

# Nombres de columna reconocidos (normalizados: minúscula, sin acentos)
_LIVA_FECHA   = {"fecha", "fecha de emision", "fecha emision"}
_LIVA_TIPO    = {"tipo", "tipo de comprobante", "tipo comprobante"}
_LIVA_PTOVTA  = {"punto de venta", "pto venta", "punto venta"}
_LIVA_NRODESDE= {"numero desde", "nro desde", "numero", "nro"}
_LIVA_TIPODOC = {"tipo doc vendedor", "tipo doc emisor", "tipo doc"}
_LIVA_NRODOC  = {"nro doc vendedor", "nro doc emisor", "nro doc", "cuit", "cuit vendedor", "cuit emisor", "nro doc vendedor emisor"}
_LIVA_DENOM   = {"denominacion vendedor", "denominacion emisor", "denominacion", "razon social", "vendedor", "emisor", "proveedor"}
_LIVA_MONEDA  = {"moneda"}
_LIVA_NETO    = {"neto gravado", "neto", "importe neto gravado", "neto gravado total"}
_LIVA_NOGRAV  = {"no gravado", "importe no gravado", "conceptos no gravado"}
_LIVA_EXENTO  = {"exento", "op exentas", "operaciones exentas", "importe exento"}
_LIVA_IVA     = {"iva", "importe iva", "total iva", "iva liquidado"}
_LIVA_TOTAL   = {"total", "importe total", "total comprobante", "imp total"}


def _norm_liva(s: str) -> str:
    """Normaliza headers del Libro IVA: minúscula, sin acentos, sin puntos,
    espacios colapsados. Tolera el mojibake Latin-1 de algunas hojas de AFIP."""
    s = (s or "").strip().lower()
    # Arreglar mojibake común (Latin-1 mal decodificado)
    for a, b in [("ã³", "o"), ("ãº", "u"), ("ã©", "e"), ("ã¡", "a"), ("ã­", "i"), ("ã±", "n")]:
        s = s.replace(a, b)
    for a, b in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("ñ","n")]:
        s = s.replace(a, b)
    s = s.replace(".", " ").replace("-", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _liva_find_col(headers_norm: list[str], names: set) -> int | None:
    """Busca el índice de columna cuyo header normalizado coincide con names."""
    for i, h in enumerate(headers_norm):
        if h in names:
            return i
    # Búsqueda parcial (por si el header trae texto extra)
    for i, h in enumerate(headers_norm):
        for n in names:
            if n and n in h:
                return i
    return None


def _liva_best_sheet(content: bytes):
    """Elige la hoja que tenga las columnas del Libro IVA (fecha + CUIT + neto/IVA)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    best = None
    best_score = -1
    for ws in wb.worksheets:
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(list(row))
            if i > 40:  # con las primeras filas alcanza para detectar el header
                break
        if not rows:
            continue
        # Buscar la fila de encabezado: la que más nombres conocidos matchea
        for hi in range(min(8, len(rows))):
            hdr = [_norm_liva(str(c)) if c is not None else "" for c in rows[hi]]
            score = 0
            for names in (_LIVA_FECHA, _LIVA_NRODOC, _LIVA_DENOM, _LIVA_NETO, _LIVA_IVA, _LIVA_TOTAL):
                if _liva_find_col(hdr, names) is not None:
                    score += 1
            if score > best_score:
                best_score = score
                best = (ws.title, hi)
    if best is None or best_score < 3:
        return None
    # Releer la hoja elegida entera
    ws = wb[best[0]]
    all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    return {"title": best[0], "header_row": best[1], "rows": all_rows}


def _parse_cuit(value) -> str | None:
    """Normaliza un CUIT a 11 dígitos (sin guiones)."""
    if value is None:
        return None
    s = re.sub(r"\D", "", str(value))
    return s if len(s) == 11 else None


def parse_libro_iva(content: bytes, filename: str = "") -> dict:
    """Parsea el Libro IVA Compras de AFIP.

    Devuelve proveedores únicos (CUIT + denominación) y comprobantes de
    compra (a pagar) con los importes discriminados: neto, no gravado,
    exento, IVA y total.
    """
    if not (filename.lower().endswith((".xlsx", ".xlsm"))):
        return {"ok": False, "error": "El Libro IVA de AFIP tiene que ser un Excel (.xlsx)."}

    try:
        sheet = _liva_best_sheet(content)
    except Exception as e:
        return {"ok": False, "error": f"No se pudo leer el Excel: {e}"}
    if not sheet:
        return {"ok": False, "error": "No reconocí el formato del Libro IVA. Verificá que sea el export de AFIP con columnas de proveedor e IVA."}

    rows = sheet["rows"]
    hi = sheet["header_row"]
    headers_norm = [_norm_liva(str(c)) if c is not None else "" for c in rows[hi]]

    col = {
        "fecha":   _liva_find_col(headers_norm, _LIVA_FECHA),
        "tipo":    _liva_find_col(headers_norm, _LIVA_TIPO),
        "ptovta":  _liva_find_col(headers_norm, _LIVA_PTOVTA),
        "nrodesde":_liva_find_col(headers_norm, _LIVA_NRODESDE),
        "cuit":    _liva_find_col(headers_norm, _LIVA_NRODOC),
        "denom":   _liva_find_col(headers_norm, _LIVA_DENOM),
        "moneda":  _liva_find_col(headers_norm, _LIVA_MONEDA),
        "neto":    _liva_find_col(headers_norm, _LIVA_NETO),
        "nograv":  _liva_find_col(headers_norm, _LIVA_NOGRAV),
        "exento":  _liva_find_col(headers_norm, _LIVA_EXENTO),
        "iva":     _liva_find_col(headers_norm, _LIVA_IVA),
        "total":   _liva_find_col(headers_norm, _LIVA_TOTAL),
    }
    if col["cuit"] is None or col["denom"] is None:
        return {"ok": False, "error": "No encontré las columnas de CUIT y proveedor en el Libro IVA."}

    def cell(row, key):
        i = col[key]
        return row[i] if (i is not None and i < len(row)) else None

    proveedores = {}   # cuit -> {cuit, nombre, comprobantes, total}
    comprobantes = []
    skipped = 0

    for row in rows[hi + 1:]:
        if not row or all(c is None for c in row):
            continue
        cuit = _parse_cuit(cell(row, "cuit"))
        denom = cell(row, "denom")
        if not cuit or not denom:
            skipped += 1
            continue
        denom = str(denom).strip()

        fecha = _parse_date(cell(row, "fecha"))
        neto = _parse_amount(cell(row, "neto")) or 0.0
        nograv = _parse_amount(cell(row, "nograv")) or 0.0
        exento = _parse_amount(cell(row, "exento")) or 0.0
        iva = _parse_amount(cell(row, "iva")) or 0.0
        total = _parse_amount(cell(row, "total"))
        if total is None:
            total = neto + nograv + exento + iva

        # Tipo y número de comprobante
        tipo = cell(row, "tipo")
        tipo_str = str(tipo).strip() if tipo is not None else ""
        pv = cell(row, "ptovta")
        nro = cell(row, "nrodesde")
        numero = ""
        if pv is not None and nro is not None:
            try:
                numero = f"{int(pv):04d}-{int(nro):08d}"
            except (ValueError, TypeError):
                numero = f"{pv}-{nro}"

        moneda_raw = str(cell(row, "moneda") or "").strip()
        moneda = "USD" if ("US" in moneda_raw.upper() or "U$" in moneda_raw) else "ARS"

        # Acumular proveedor
        if cuit not in proveedores:
            proveedores[cuit] = {"cuit": cuit, "nombre": denom, "comprobantes": 0, "total": 0.0}
        proveedores[cuit]["comprobantes"] += 1
        proveedores[cuit]["total"] += (total or 0.0)

        comprobantes.append({
            "tipo": "pagar",
            "contraparte": denom,
            "cuit": cuit,
            "numero": numero,
            "tipoComprobante": tipo_str,
            "emision": fecha,
            "moneda": moneda,
            "neto": round(neto, 2),
            "noGravado": round(nograv, 2),
            "exento": round(exento, 2),
            "iva": round(iva, 2),
            "monto": round(total or 0.0, 2),
        })

    # Formatear CUIT con guiones para mostrar
    def fmt_cuit(c):
        return f"{c[:2]}-{c[2:10]}-{c[10:]}" if len(c) == 11 else c

    prov_list = sorted(proveedores.values(), key=lambda p: -p["total"])
    for p in prov_list:
        p["cuitFmt"] = fmt_cuit(p["cuit"])
        p["total"] = round(p["total"], 2)

    total_iva = round(sum(c["iva"] for c in comprobantes), 2)
    total_neto = round(sum(c["neto"] for c in comprobantes), 2)
    total_monto = round(sum(c["monto"] for c in comprobantes), 2)

    return {
        "ok": True,
        "proveedores": prov_list,
        "comprobantes": comprobantes,
        "resumen": {
            "proveedores": len(prov_list),
            "comprobantes": len(comprobantes),
            "total_neto": total_neto,
            "total_iva": total_iva,
            "total": total_monto,
            "skipped": skipped,
            "sheet": sheet["title"],
        },
    }


# ══════════════════════════════════════════════════════════════════
# Import de Retenciones y Percepciones de AFIP ("Mis Retenciones")
# ══════════════════════════════════════════════════════════════════
# Los tres exports (SICORE, Aduana, Ganancias) comparten el mismo
# formato: CUIT del agente, impuesto, régimen, fecha, tipo
# (RETENCION/PERCEPCION) e importe. Un solo parser los cubre.

def _rows_from_xls(content: bytes) -> list[list]:
    """Lee un .xls binario viejo (formato AFIP) con xlrd."""
    try:
        import xlrd
    except ImportError:
        raise RuntimeError("xlrd no está instalado (necesario para .xls). Ejecutá: pip install xlrd")
    wb = xlrd.open_workbook(file_contents=content)
    sh = wb.sheet_by_index(0)
    return [sh.row_values(i) for i in range(sh.nrows)]


# Columnas de "Mis Retenciones" (normalizadas con _norm_liva)
_RET_CUIT   = {"cuit agente ret per", "cuit agente", "cuit agente ret perc", "cuit"}
_RET_DENOM  = {"denominacion o razon social", "denominacion o razon", "denominacion", "razon social", "agente"}
_RET_IMP    = {"descripcion impuesto", "impuesto desc", "descripcion imp"}
_RET_IMPCOD = {"impuesto"}
_RET_REGDESC= {"descripcion regimen", "regimen desc", "descripcion regim"}
_RET_FECHA  = {"fecha ret perc", "fecha ret perc", "fecha", "fecha retencion"}
_RET_TIPO   = {"descripcion operacion", "operacion", "tipo"}
_RET_IMPORTE= {"importe ret perc", "importe", "importe retencion"}
_RET_CERT   = {"numero certificado", "certificado", "nro certificado"}


def parse_retenciones(content: bytes, filename: str = "") -> dict:
    """Parsea un export de 'Mis Retenciones y Percepciones' de AFIP.

    Funciona con SICORE, Aduana y Ganancias (mismo formato). Acepta .xls
    (formato viejo) y .xlsx. Devuelve las retenciones/percepciones con
    agente, impuesto, fecha, tipo (retención o percepción) e importe.
    """
    fn = filename.lower()
    try:
        if fn.endswith(".xls"):
            rows = _rows_from_xls(content)
        elif fn.endswith((".xlsx", ".xlsm")):
            # Puede tener varias hojas; buscar la que tenga los datos
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            rows = None
            for ws in wb.worksheets:
                first = []
                for i, r in enumerate(ws.iter_rows(values_only=True)):
                    first.append(list(r))
                    if i > 2: break
                # ¿El header tiene "cuit agente" e "importe"?
                if first:
                    hdr = [_norm_liva(str(c)) if c is not None else "" for c in first[0]]
                    if _liva_find_col(hdr, _RET_CUIT) is not None and _liva_find_col(hdr, _RET_IMPORTE) is not None:
                        rows = [list(r) for r in ws.iter_rows(values_only=True)]
                        break
            if rows is None:
                return {"ok": False, "error": "No encontré la hoja de retenciones en el Excel."}
        else:
            return {"ok": False, "error": "El archivo tiene que ser .xls o .xlsx de AFIP."}
    except Exception as e:
        return {"ok": False, "error": f"No se pudo leer el archivo: {e}"}

    if not rows:
        return {"ok": False, "error": "El archivo está vacío."}

    hdr = [_norm_liva(str(c)) if c is not None else "" for c in rows[0]]
    col = {
        "cuit":    _liva_find_col(hdr, _RET_CUIT),
        "denom":   _liva_find_col(hdr, _RET_DENOM),
        "imp":     _liva_find_col(hdr, _RET_IMP),
        "regdesc": _liva_find_col(hdr, _RET_REGDESC),
        "fecha":   _liva_find_col(hdr, _RET_FECHA),
        "tipo":    _liva_find_col(hdr, _RET_TIPO),
        "importe": _liva_find_col(hdr, _RET_IMPORTE),
        "cert":    _liva_find_col(hdr, _RET_CERT),
    }
    if col["cuit"] is None or col["importe"] is None:
        return {"ok": False, "error": "No reconocí el formato de 'Mis Retenciones' de AFIP (faltan CUIT o importe)."}

    def cell(row, key):
        i = col[key]
        return row[i] if (i is not None and i < len(row)) else None

    items = []
    tot_ret = tot_per = 0.0
    for row in rows[1:]:
        if not row or all((c is None or str(c).strip() == "") for c in row):
            continue
        cuit = _parse_cuit(cell(row, "cuit"))
        importe = _parse_amount(cell(row, "importe"))
        if importe is None or importe == 0:
            continue
        tipo_raw = str(cell(row, "tipo") or "").strip().upper()
        es_percep = "PERCEP" in tipo_raw
        tipo = "percepcion" if es_percep else "retencion"
        if es_percep:
            tot_per += importe
        else:
            tot_ret += importe
        items.append({
            "cuit": cuit,
            "agente": str(cell(row, "denom") or "").strip(),
            "impuesto": str(cell(row, "imp") or "").strip(),
            "regimen": str(cell(row, "regdesc") or "").strip(),
            "fecha": _parse_date(cell(row, "fecha")),
            "tipo": tipo,
            "importe": round(importe, 2),
            "certificado": str(cell(row, "cert") or "").strip(),
        })

    # Agrupar por impuesto para el resumen
    por_impuesto = {}
    for it in items:
        k = it["impuesto"] or "(sin descripción)"
        if k not in por_impuesto:
            por_impuesto[k] = {"impuesto": k, "tipo": it["tipo"], "cantidad": 0, "total": 0.0}
        por_impuesto[k]["cantidad"] += 1
        por_impuesto[k]["total"] += it["importe"]
    resumen_imp = sorted(por_impuesto.values(), key=lambda x: -x["total"])
    for r in resumen_imp:
        r["total"] = round(r["total"], 2)

    return {
        "ok": True,
        "items": items,
        "por_impuesto": resumen_imp,
        "resumen": {
            "cantidad": len(items),
            "total_retenciones": round(tot_ret, 2),
            "total_percepciones": round(tot_per, 2),
            "total": round(tot_ret + tot_per, 2),
        },
    }
